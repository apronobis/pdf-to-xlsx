import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font


def write_to_xlsx(data):
    wb = openpyxl.Workbook()
    ws = wb.active

    header = data["ID"]
    col_index = [2]
    for title, fields in header.items():
        ws.cell(row=1, column=col_index[-1], value=title)
        col_index.append(col_index[-1] + len(fields))
        ws.merge_cells(
            start_row=1,
            end_row=1,
            start_column=col_index[-2],
            end_column=col_index[-1] - 1,
        )

    for _id, window in data.items():
        row = [_id]
        for title, fields in header.items():
            if title not in window:
                row.extend([""] * len(fields))
            else:
                row.extend(window[title])
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[1].column_letter].width = len(col[1].value) + 2

    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 16

    for row in ws.iter_rows(min_row=2):
        for i in col_index:
            row[i - 2].border = Border(right=Side(style="thick"))

    for cell in ws[1] + ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    wb.save("output/save.xlsx")

    return wb, ws
